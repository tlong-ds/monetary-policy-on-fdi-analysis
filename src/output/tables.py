"""src/output/tables.py

Regression table formatting and export (instruct.md §13 outputs #7–8).

Supports:
  - Static results (linearmodels PanelOLS / RandomEffects / PooledOLS)
  - Dynamic results (System GMM dict from dynamic.py)

Output formats:
  - Excel (.xlsx) with formatted cells
  - LaTeX (.tex) tabular environment

Table structure (per column = one model spec):
  Row per variable: coefficient (with stars) | SE in parentheses
  Footer rows:      N, R² (within), Country FE, Year FE, Estimator,
                    [GMM only]: N instruments, Hansen p, AR(1) p, AR(2) p
"""
from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import pandas as pd

from src.estimation.specs import EXPECTED_SIGNS, VARIABLE_LABELS


# ---------------------------------------------------------------------------
# Significance stars
# ---------------------------------------------------------------------------
def _stars(pval: float) -> str:
    if np.isnan(pval):
        return ""
    if pval < 0.01:
        return "***"
    if pval < 0.05:
        return "**"
    if pval < 0.10:
        return "*"
    return ""


def _fmt_coef(coef: float, pval: float, digits: int = 3) -> str:
    if np.isnan(coef):
        return ""
    return f"{coef:.{digits}f}{_stars(pval)}"


def _fmt_se(se: float, digits: int = 3) -> str:
    if np.isnan(se):
        return ""
    return f"({se:.{digits}f})"


# ---------------------------------------------------------------------------
# Extract params/SE/p from a result object
# (handles both linearmodels results and the dynamic GMM dict)
# ---------------------------------------------------------------------------
def _extract_coef_table(result_dict: dict) -> pd.DataFrame:
    """Return tidy DataFrame: variable, coef, se, pval."""
    res = result_dict.get("preferred_result")

    if res is not None:
        # linearmodels result object
        try:
            params  = res.params
            se      = res.std_errors
            pvals   = res.pvalues
        except AttributeError:
            params  = res.params
            se      = res.bse
            pvals   = res.pvalues

        rows = []
        for var in params.index:
            rows.append({
                "variable": var,
                "coef": float(params[var]),
                "se":   float(se[var]),
                "pval": float(pvals[var]),
            })
        return pd.DataFrame(rows)

    # Dynamic GMM dict
    if "params" in result_dict:
        rows = []
        for var in result_dict["params"].index:
            rows.append({
                "variable": var,
                "coef": float(result_dict["params"][var]),
                "se":   float(result_dict["std_errors"][var]),
                "pval": float(result_dict["p_values"][var]),
            })
        return pd.DataFrame(rows)

    return pd.DataFrame(columns=["variable", "coef", "se", "pval"])


# ---------------------------------------------------------------------------
# Variable order for table rows (follows instruct.md §4 ordering)
# ---------------------------------------------------------------------------
VARIABLE_ORDER: list[str] = [
    "fdi_pct_gdp_lag1",
    "bm_growth",
    "real_interest_rate",
    "gdp_growth",
    "ln_gdppc",
    "trade_pct_gdp",
    "d_ln_exr",
    "inflation",
    "rq",
    "rl",
    "wgi_composite",
]


def _model_footer(result_dict: dict) -> dict[str, str]:
    """Extract footer statistics for one model column."""
    footer: dict[str, str] = {}

    res = result_dict.get("preferred_result")
    est = result_dict.get("estimator_label", "")

    if res is not None:
        footer["N"] = str(result_dict.get("obs", getattr(res, "nobs", "")))
        # R²
        try:
            rsq = res.rsquared_within if hasattr(res, "rsquared_within") else res.rsquared
            footer["R² (within)"] = f"{float(rsq):.3f}" if not np.isnan(float(rsq)) else "—"
        except Exception:
            footer["R² (within)"] = "—"

        footer["Country FE"] = "Yes" if "Fixed" in est else ("—" if "Pooled" in est else "Yes")
        footer["Year FE"]    = "Yes"
        footer["Estimator"]  = est

    else:
        # GMM
        footer["N"]           = str(result_dict.get("n_obs", ""))
        footer["R² (within)"] = "—"
        footer["Country FE"]  = "—"
        footer["Year FE"]     = "Yes"
        footer["Estimator"]   = result_dict.get("estimator_label", "System GMM")
        footer["N instruments"] = str(result_dict.get("n_instruments", ""))
        h = result_dict.get("hansen", {})
        footer["Hansen p-value"] = f"{h.get('p_value', '')}" if "p_value" in h else "—"
        ar1 = result_dict.get("ar1", {})
        ar2 = result_dict.get("ar2", {})
        footer["AR(1) p-value"] = f"{ar1.get('p_value', '')}"
        footer["AR(2) p-value"] = f"{ar2.get('p_value', '')}"

    return footer


# ---------------------------------------------------------------------------
# Build the combined wide table (variables × specs)
# ---------------------------------------------------------------------------
def build_regression_table(
    results: list[dict],
    variable_order: list[str] | None = None,
) -> pd.DataFrame:
    """Build a wide regression table (rows = variables/stats, cols = specs).

    Each variable has two rows: coefficient and (SE).
    """
    var_order = variable_order or VARIABLE_ORDER

    # Collect all unique variables across models
    all_vars_seen: list[str] = []
    coef_tables: list[pd.DataFrame] = []
    footers: list[dict] = []
    col_headers: list[str] = []

    for rd in results:
        ct = _extract_coef_table(rd)
        coef_tables.append(ct)
        footers.append(_model_footer(rd))
        col_headers.append(f"{rd.get('spec_label','')}\n{rd.get('spec_name','')}")
        for v in ct["variable"].tolist():
            if v not in all_vars_seen:
                all_vars_seen.append(v)

    # Order: use var_order first, then any remaining
    ordered_vars = [v for v in var_order if v in all_vars_seen]
    remaining = [v for v in all_vars_seen if v not in ordered_vars
                 and not v.startswith("yr_") and v not in ("Intercept", "const", "Intercept")]
    display_vars = ordered_vars + remaining

    # Collect all footer keys
    all_footer_keys: list[str] = []
    for f in footers:
        for k in f:
            if k not in all_footer_keys:
                all_footer_keys.append(k)

    # Build rows
    rows: list[dict] = []
    for var in display_vars:
        label = VARIABLE_LABELS.get(var, var)
        row_coef: dict = {"variable": label, "row_type": "coef"}
        row_se:   dict = {"variable": "",    "row_type": "se"}

        for col_h, ct in zip(col_headers, coef_tables):
            match = ct[ct["variable"] == var]
            if match.empty:
                row_coef[col_h] = ""
                row_se[col_h]   = ""
            else:
                r = match.iloc[0]
                row_coef[col_h] = _fmt_coef(r["coef"], r["pval"])
                row_se[col_h]   = _fmt_se(r["se"])
        rows.extend([row_coef, row_se])

    # Footer rows
    for key in all_footer_keys:
        row: dict = {"variable": key, "row_type": "stat"}
        for col_h, ft in zip(col_headers, footers):
            row[col_h] = ft.get(key, "")
        rows.append(row)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def export_excel(
    table: pd.DataFrame,
    path: Path,
    sheet_name: str = "Regression Results",
    note: str = "Note: *p<0.10, **p<0.05, ***p<0.01. SE in parentheses. Cluster-robust SE at country level.",
) -> None:
    """Write regression table to Excel with basic formatting."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Main table
        display = table.drop(columns=["row_type"], errors="ignore")
        display.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]
        # Note row
        note_row = len(display) + 3
        ws.cell(row=note_row, column=1, value=note)

        # Bold header row
        from openpyxl.styles import Font, PatternFill, Alignment
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True)

        # Shade stat rows
        grey_fill = PatternFill("solid", fgColor="F2F2F2")
        for row_idx, row_type in enumerate(table["row_type"].tolist(), start=2):
            if row_type == "stat":
                for cell in ws[row_idx]:
                    cell.fill = grey_fill

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)


# ---------------------------------------------------------------------------
# LaTeX export
# ---------------------------------------------------------------------------
def _latex_escape(s: str) -> str:
    return str(s).replace("&", "\\&").replace("%", "\\%").replace("_", "\\_")


def export_latex(
    table: pd.DataFrame,
    path: Path,
    caption: str = "Panel Regression Results",
    label: str = "tab:regression",
    note: str = "\\textit{Note:} $^{*}$p$<$0.10, $^{**}$p$<$0.05, $^{***}$p$<$0.01. "
                "Standard errors (cluster-robust, country level) in parentheses.",
) -> None:
    """Write regression table as LaTeX tabular."""
    display = table.drop(columns=["row_type"], errors="ignore")
    n_cols = len(display.columns)
    col_fmt = "l" + "c" * (n_cols - 1)

    lines = [
        "\\begin{table}[htbp]",
        "\\centering",
        f"\\caption{{{_latex_escape(caption)}}}",
        f"\\label{{{label}}}",
        f"\\begin{{tabular}}{{{col_fmt}}}",
        "\\hline\\hline",
    ]

    # Header
    header = " & ".join(_latex_escape(str(c)).replace("\n", " / ") for c in display.columns)
    lines.append(header + " \\\\")
    lines.append("\\hline")

    # Body
    prev_type = None
    for idx, row in table.iterrows():
        row_type = row.get("row_type", "coef")
        vals = display.loc[idx]

        if row_type == "stat" and prev_type != "stat":
            lines.append("\\hline")

        cells = " & ".join(_latex_escape(str(v)) for v in vals.values)
        lines.append(cells + " \\\\")
        prev_type = row_type

    lines += [
        "\\hline\\hline",
        f"\\multicolumn{{{n_cols}}}{{l}}{{{note}}}",
        "\\end{tabular}",
        "\\end{table}",
    ]

    path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Convenience: save both formats
# ---------------------------------------------------------------------------
def save_regression_tables(
    results: list[dict],
    tables_dir: Path,
    stem: str = "regression_results",
    caption: str = "FDI Determinants: Panel Regression Results",
) -> None:
    """Build table and save as .xlsx and .tex."""
    table = build_regression_table(results)
    xlsx_path = tables_dir / f"{stem}.xlsx"
    tex_path  = tables_dir / f"{stem}.tex"
    export_excel(table, xlsx_path)
    export_latex(table, tex_path, caption=caption, label=f"tab:{stem}")
    print(f"  → {xlsx_path}")
    print(f"  → {tex_path}")


# ---------------------------------------------------------------------------
# Estimator selection log table
# ---------------------------------------------------------------------------
def build_selection_log(results: list[dict]) -> pd.DataFrame:
    """Summarise BP-LM and Hausman decisions per spec."""
    rows = []
    for rd in results:
        bp = rd.get("bp_lm", {})
        h  = rd.get("hausman", {})
        rows.append({
            "Spec":             rd.get("spec_label", ""),
            "Name":             rd.get("spec_name", ""),
            "BP-LM stat":       bp.get("statistic", ""),
            "BP-LM p":          bp.get("p_value", ""),
            "BP-LM decision":   bp.get("decision", ""),
            "Hausman stat":     h.get("statistic", "") if h else "—",
            "Hausman p":        h.get("p_value", "")   if h else "—",
            "Hausman decision": h.get("decision", "")   if h else "—",
            "Preferred":        rd.get("estimator_label", ""),
            "N obs":            rd.get("obs", ""),
        })
    return pd.DataFrame(rows)
